"""Applications router."""
from __future__ import annotations

import csv
import io
import sqlite3
from datetime import date
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from starlette.responses import StreamingResponse

from ... import db
from ..deps import get_db
from ..schemas import (
    ApplicationsListResponse,
    CreateApplicationRequest,
    ImportApplicationsResponse,
    SuccessResponse,
    UpdateApplicationRequest,
    UpdateApplicationStatusRequest,
)

router = APIRouter(prefix="/api/applications", tags=["applications"])
PER_PAGE = 50


@router.get("/export")
def export_applications(
    format: str = Query("csv", pattern="^(csv|tsv|excel|xlsx)$"),
    status: str | None = Query(None),
    q: str | None = Query(None),
    sort: str | None = Query(None),
    order: str | None = Query(None),
    conn: sqlite3.Connection = Depends(get_db),
):
    # Fetch all matching applications without pagination limits
    rows = db.list_applications(
        conn,
        limit=10000,
        offset=0,
        status=status,
        q=q,
        sort=sort,
        order=order,
    )

    headers = [
        "Applied Date",
        "Role Title",
        "Company",
        "Location",
        "Status",
        "Salary Entered",
        "Job URL",
    ]

    # Native Excel format (.xlsx)
    if format in ("excel", "xlsx"):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="171717", end_color="171717", fill_type="solid")
        ws.append(headers)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        for r in rows:
            d = dict(r)
            ws.append([
                d.get("applied_at") or "",
                d.get("title") or "",
                d.get("company") or "",
                d.get("location") or "",
                d.get("status") or "Submitted",
                d.get("salary_entered") or "",
                d.get("url") or "",
            ])

        # Add In-Cell Dropdown Data Validation for Status Column (Column E / 5)
        status_options = '"Submitted,Process,Interview,Offering,Declined,Rejected"'
        dv = DataValidation(
            type="list",
            formula1=status_options,
            allow_blank=True,
            showDropDown=False,  # in openpyxl, showDropDown=False enables the in-cell dropdown arrow
            showErrorMessage=True,
            errorTitle="Invalid Status",
            error="Please select a valid status from the dropdown list.",
        )
        ws.add_data_validation(dv)
        max_row = max(len(rows) + 100, 200)
        dv.add(f"E2:E{max_row}")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

        out_stream = io.BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        return Response(
            content=out_stream.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="applications.xlsx"'},
        )

    output = io.StringIO()
    # Write UTF-8 BOM for Excel compatibility
    output.write("\ufeff")

    delimiter = "\t" if format == "tsv" else ","
    writer = csv.writer(output, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)

    for r in rows:
        d = dict(r)
        writer.writerow([
            d.get("applied_at") or "",
            d.get("title") or "",
            d.get("company") or "",
            d.get("location") or "",
            d.get("status") or "Submitted",
            d.get("salary_entered") or "",
            d.get("url") or "",
        ])

    data_bytes = output.getvalue().encode("utf-8-sig")

    if format == "tsv":
        filename = "applications.tsv"
        media_type = "text/tab-separated-values; charset=utf-8"
    else:
        filename = "applications.csv"
        media_type = "text/csv; charset=utf-8"

    return Response(
        content=data_bytes,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("", response_model=ApplicationsListResponse)
def list_applications(
    page: int = Query(1, ge=1),
    per_page: int = Query(PER_PAGE, ge=1, le=200),
    status: str | None = Query(None),
    q: str | None = Query(None),
    sort: str | None = Query(None),
    order: str | None = Query(None),
    conn: sqlite3.Connection = Depends(get_db),
):
    rows = db.list_applications(
        conn,
        limit=per_page + 1,
        offset=(page - 1) * per_page,
        status=status,
        q=q,
        sort=sort,
        order=order,
    )
    has_next = len(rows) > per_page
    total = db.count_applications(conn, status=status, q=q)
    return {
        "apps": [dict(r) for r in rows[:per_page]],
        "page": page,
        "has_next": has_next,
        "total": total,
    }


@router.post("", response_model=SuccessResponse)
def create_application(
    payload: CreateApplicationRequest,
    conn: sqlite3.Connection = Depends(get_db),
):
    """Manually add a new application."""
    data = payload.model_dump()
    app_id = db.create_manual_application(conn, data)
    return SuccessResponse(message=f"Application created successfully with ID {app_id}")


@router.put("/{app_id}", response_model=SuccessResponse)
def update_application_endpoint(
    app_id: int,
    payload: UpdateApplicationRequest,
    conn: sqlite3.Connection = Depends(get_db),
):
    """Update fields of an application and its linked job."""
    data = payload.model_dump(exclude_unset=True)
    updated = db.update_application(conn, app_id, data)
    if not updated:
        raise HTTPException(status_code=404, detail="Application not found")
    return SuccessResponse(message="Application updated successfully")


@router.delete("/{app_id}", response_model=SuccessResponse)
def delete_application_endpoint(
    app_id: int,
    conn: sqlite3.Connection = Depends(get_db),
):
    """Delete an application."""
    deleted = db.delete_application(conn, app_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Application not found")
    return SuccessResponse(message="Application deleted successfully")


@router.post("/import", response_model=ImportApplicationsResponse)
async def import_applications(
    file: UploadFile = File(...),
    conn: sqlite3.Connection = Depends(get_db),
):
    """Import applications from a CSV, TSV, or XLSX file."""
    filename = (file.filename or "").lower()
    content = await file.read()

    rows_to_process: list[dict[str, str]] = []

    if filename.endswith(".xls") and not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Legacy Excel (.xls) format is not supported. Please use .xlsx or .csv.")

    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            if ws is None:
                raise ValueError("Workbook contains no active sheet")

            iter_rows = ws.iter_rows(values_only=True)
            header_row = next(iter_rows, None)
            if not header_row:
                raise ValueError("Spreadsheet is empty")

            col_names = [str(c or "").strip() for c in header_row]
            for row in iter_rows:
                if not any(row):
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(col_names):
                        row_dict[col_names[idx]] = str(val or "").strip()
                rows_to_process.append(row_dict)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")
    else:
        # CSV or TSV
        try:
            # Handle BOM and decode
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("latin-1")
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Failed to decode file: {e}")

        lines = text.splitlines()
        first_line = lines[0] if lines else ""
        delimiter = "\t" if filename.endswith(".tsv") or "\t" in first_line else ","
        try:
            reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            for row in reader:
                # Strip all keys and values
                cleaned = {str(k or "").strip(): str(v or "").strip() for k, v in row.items()}
                if any(cleaned.values()):
                    rows_to_process.append(cleaned)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV/TSV: {e}")

    if not rows_to_process:
        return ImportApplicationsResponse(
            success=True,
            imported=0,
            skipped=0,
            errors=[],
            message="No rows found in file",
        )

    # Validate column mapping
    # Expected columns: Applied Date, Role Title, Company, Location, Status, Salary Entered, Job URL
    imported = 0
    skipped = 0
    errors: list[str] = []

    today_str = date.today().isoformat()

    for idx, r in enumerate(rows_to_process, 1):
        # Extract fields matching export headers
        title = r.get("Role Title") or r.get("Role") or r.get("title") or ""
        company = r.get("Company") or r.get("company") or ""
        url = r.get("Job URL") or r.get("URL") or r.get("url") or ""
        applied_at = r.get("Applied Date") or r.get("Applied") or r.get("applied_at") or today_str
        location = r.get("Location") or r.get("location") or ""
        status = r.get("Status") or r.get("status") or "Submitted"
        salary_entered = r.get("Salary Entered") or r.get("Salary") or r.get("salary_entered") or ""

        # Validate required columns: title, company, url
        if not title or not company or not url:
            errors.append(
                f"Row {idx}: missing required field(s) (Title: '{title}', Company: '{company}', URL: '{url}')"
            )
            continue

        # Check for existing duplicate by (title, company, url)
        existing = db.find_application_by_job_details(conn, title, company, url)
        if existing:
            skipped += 1
            continue

        try:
            db.create_manual_application(
                conn,
                {
                    "title": title,
                    "company": company,
                    "url": url,
                    "applied_at": applied_at,
                    "location": location,
                    "status": status,
                    "salary_entered": salary_entered,
                },
            )
            imported += 1
        except Exception as e:
            errors.append(f"Row {idx}: failed to insert ({e})")

    return ImportApplicationsResponse(
        success=True,
        imported=imported,
        skipped=skipped,
        errors=errors,
        message=f"Imported {imported} row(s), skipped {skipped} duplicate(s)"
        + (f" with {len(errors)} error(s)" if errors else ""),
    )


@router.patch("/{app_id}/status", response_model=SuccessResponse)
def update_status(
    app_id: int,
    payload: UpdateApplicationStatusRequest,
    conn: sqlite3.Connection = Depends(get_db),
):
    updated = db.update_application_status(conn, app_id, payload.status)
    if not updated:
        raise HTTPException(status_code=404, detail="Application not found")
    return SuccessResponse(message="Status updated successfully")

